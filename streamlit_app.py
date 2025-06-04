#! streamlit run streamlit_app.py --server.port=8501 --server.enableCORS=false


import streamlit as st
import pandas as pd
from income import get_income_by_name  # 너의 함수 불러오기
import tempfile
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


st.title("📊 DART 재무제표 추출기")

corp_name = st.text_input("기업명", "삼성전자")
corp_market = st.text_input("기업 구분 ('Y': 코스피, 'K': 코스닥, 'N': 코넥스, 'E': 기타)", "Y")
bgn_de = st.text_input("시작일", "20220101")
end_de = st.text_input("종료일", "20241231")

if st.button("📥 엑셀 파일 생성 및 다운로드"):

    progress_bar = st.progress(0)
    status_text = st.empty()

    def update_progress(pct, message):
        progress_bar.progress(pct)
        status_text.text(message)


    # 🔽 너의 노트북에 있는 처리 함수 호출
    dfs = get_income_by_name(  # 이건 너가 만든 함수
        corp_name=corp_name,
        corp_market=corp_market,
        bgn_de=bgn_de,
        end_de=end_de,
        progress_fn = update_progress
    )


    # 임시 엑셀 파일 생성
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        filepath = tmp.name

    with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
        for sheet_name, df in dfs:
            df.to_excel(writer, sheet_name=sheet_name, index=False)

    # 엑셀 파일 열어서 서식 적용
    wb = load_workbook(filepath)

    for sheetname in wb.sheetnames:
        ws = wb[sheetname]
        ws.column_dimensions['A'].width = 40  # label_ko 열

        for col in range(2, ws.max_column + 1):
            col_letter = get_column_letter(col)
            max_width = 10
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=col)
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0'
                    formatted = f"{cell.value:,.0f}"
                    max_width = max(max_width, len(formatted) + 2)
            ws.column_dimensions[col_letter].width = max_width

    wb.save(filepath)

    # 다운로드 버튼 제공
    with open(filepath, 'rb') as f:
        st.download_button(
            label="📥 엑셀 다운로드",
            data=f,
            file_name=f"{corp_name}_포괄손익계산서.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )